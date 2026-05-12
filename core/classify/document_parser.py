
import io


def extract_text_from_bytes(data: bytes, filename: str) -> str:
    filename = (filename or "").lower()

    try:
        if filename.endswith(".pdf"):
            text = _extract_pdf(data)
            if text.strip():
                return text
            return _ocr_pdf(data)

        if filename.endswith(".docx"):
            return _extract_docx(data)

        if filename.endswith(".xlsx"):
            return _extract_xlsx(data)

        if filename.endswith((".png", ".jpg", ".jpeg", ".tiff", ".bmp")):
            return _ocr_image(data)

        if filename.endswith(".txt"):
            return data.decode("utf-8", errors="ignore")

        return data.decode("utf-8", errors="ignore")

    except Exception as e:
        print(f"⚠️ Document parse failed ({filename}): {e}")
        return ""


def _extract_pdf(data: bytes) -> str:
    import pdfplumber

    text_parts: list[str] = []
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                text_parts.append(t)
    return "\n".join(text_parts)


def _ocr_pdf(data: bytes) -> str:
    try:
        from pdf2image import convert_from_bytes
        import pytesseract
    except Exception as e:
        print(f"⚠️ OCR PDF dependencies missing: {e}")
        return ""

    text_parts: list[str] = []
    images = convert_from_bytes(data)
    for img in images:
        t = pytesseract.image_to_string(img)
        if t:
            text_parts.append(t)
    return "\n".join(text_parts)


def _extract_docx(data: bytes) -> str:
    from docx import Document

    doc = Document(io.BytesIO(data))
    return "\n".join(p.text for p in doc.paragraphs if p.text)


def _extract_xlsx(data: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), data_only=True)
    text_parts: list[str] = []

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    text_parts.append(str(cell))

    return "\n".join(text_parts)


def _ocr_image(data: bytes) -> str:
    try:
        from PIL import Image
        import pytesseract
    except Exception as e:
        print(f"⚠️ OCR image dependencies missing: {e}")
        return ""

    img = Image.open(io.BytesIO(data))
    return pytesseract.image_to_string(img)


